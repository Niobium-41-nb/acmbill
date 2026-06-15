import os
import uuid
import io
import zipfile
import tempfile
import shutil
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app, send_from_directory, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from models import db, User, Project, TeamMember, ReimbursementItem, ProjectFile, EmailVerificationCode

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

project_bp = Blueprint('project', __name__, url_prefix='/project')


def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in current_app.config.get('ALLOWED_EXTENSIONS', {})


def save_uploaded_file(file, project_id, file_type, user_id, item_id=None):
    """保存上传的文件"""
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        # 生成唯一文件名
        unique_name = f"{uuid.uuid4().hex}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{ext}"
        upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project_id))
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, unique_name)
        file.save(file_path)

        project_file = ProjectFile(
            project_id=project_id,
            item_id=item_id,
            file_type=file_type,
            original_filename=file.filename,
            stored_filename=unique_name,
            file_size=os.path.getsize(file_path),
            uploaded_by=user_id
        )
        db.session.add(project_file)
        db.session.commit()
        return project_file
    return None


def check_project_access(project):
    """检查当前用户是否有权访问该项目"""
    if current_user.is_admin():
        return True
    if project.owner_id == current_user.id:
        return True
    # 检查是否为团队成员
    return TeamMember.query.filter_by(
        project_id=project.id, user_id=current_user.id
    ).first() is not None


@project_bp.route('/dashboard')
@login_required
def dashboard():
    """项目仪表盘"""
    if current_user.is_admin():
        projects = Project.query.order_by(Project.updated_at.desc()).all()
    else:
        # 用户创建的项目 + 作为团队成员的项目
        owned = Project.query.filter_by(owner_id=current_user.id)
        team_project_ids = [tm.project_id for tm in
                           TeamMember.query.filter_by(user_id=current_user.id).all()]
        team_projects = Project.query.filter(Project.id.in_(team_project_ids)) if team_project_ids else []
        from sqlalchemy import union
        if team_project_ids:
            projects = owned.union(team_projects).order_by(Project.updated_at.desc()).all()
        else:
            projects = owned.order_by(Project.updated_at.desc()).all()
    return render_template('dashboard.html', projects=projects)


@project_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_project():
    """创建新项目"""
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        competition_name = request.form.get('competition_name', '').strip()
        competition_date = request.form.get('competition_date', '').strip()
        teacher = request.form.get('teacher', '').strip()

        if not title:
            flash('请输入项目标题', 'danger')
            return render_template('project_form.html', project=None)

        project = Project(
            title=title,
            description=description,
            competition_name=competition_name,
            competition_date=competition_date,
            owner_id=current_user.id
        )
        db.session.add(project)
        db.session.commit()
        flash('项目创建成功', 'success')
        return redirect(url_for('project.project_detail', project_id=project.id))

    return render_template('project_form.html', project=None)


@project_bp.route('/<int:project_id>')
@login_required
def project_detail(project_id):
    """项目详情页"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))

    items = project.reimbursement_items.order_by(ReimbursementItem.created_at).all()
    members = project.team_members.all()
    # 所有文件（包括项目级和条目级）
    all_files = ProjectFile.query.filter_by(project_id=project_id).order_by(ProjectFile.created_at.desc()).all()
    # 项目级文件（未关联条目的）
    files = ProjectFile.query.filter_by(project_id=project_id, item_id=None).order_by(ProjectFile.created_at.desc()).all()

    # 为每个条目统计文件
    item_files = {}
    for item in items:
        inv_count = ProjectFile.query.filter_by(project_id=project_id, item_id=item.id, file_type='invoice').count()
        pay_count = ProjectFile.query.filter_by(project_id=project_id, item_id=item.id, file_type='payment').count()
        item_files[item.id] = {'invoice': inv_count, 'payment': pay_count}

    total_amount = sum(item.amount for item in items if item.amount)

    return render_template('project_detail.html',
                           project=project, items=items, members=members,
                           files=files, all_files=all_files,
                           item_files=item_files,
                           total_amount=total_amount)


@project_bp.route('/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    """编辑项目"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))

    if request.method == 'POST':
        project.title = request.form.get('title', '').strip()
        project.description = request.form.get('description', '').strip()
        project.competition_name = request.form.get('competition_name', '').strip()
        project.competition_date = request.form.get('competition_date', '').strip()
        project.teacher = request.form.get('teacher', '').strip()
        db.session.commit()
        flash('项目信息已更新', 'success')
        return redirect(url_for('project.project_detail', project_id=project.id))

    return render_template('project_form.html', project=project)


@project_bp.route('/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    """删除项目（仅创建者或管理员）"""
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id and not current_user.is_admin():
        flash('无权删除', 'danger')
        return redirect(url_for('project.dashboard'))

    # 删除上传的文件
    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project_id))
    if os.path.exists(upload_dir):
        shutil.rmtree(upload_dir)

    db.session.delete(project)
    db.session.commit()
    flash('项目已删除', 'success')
    return redirect(url_for('project.dashboard'))


@project_bp.route('/<int:project_id>/status', methods=['POST'])
@login_required
def update_status(project_id):
    """更新项目状态"""
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id and not current_user.is_admin():
        flash('无权操作', 'danger')
        return redirect(url_for('project.dashboard'))

    new_status = request.form.get('status', '')
    if new_status in ['draft', 'submitted', 'approved', 'rejected']:
        project.status = new_status
        db.session.commit()
        flash('状态已更新', 'success')
    return redirect(url_for('project.project_detail', project_id=project_id))


@project_bp.route('/<int:project_id>/reimburse-status', methods=['POST'])
@login_required
def update_reimburse_status(project_id):
    """更新报销状态（仅管理员可操作）"""
    project = Project.query.get_or_404(project_id)
    if not current_user.is_admin():
        flash('无权操作', 'danger')
        return redirect(url_for('project.dashboard'))

    new_status = request.form.get('reimburse_status', '')
    if new_status in ['draft', 'cannot_insure', 'reimbursed']:
        project.reimburse_status = new_status
        db.session.commit()
        flash('报销状态已更新', 'success')
    else:
        flash('无效的状态值', 'danger')
    return redirect(url_for('project.project_detail', project_id=project_id))


def check_item_file_requirement(project_id, expense_category, item_id=None):
    """检查条目是否满足文件要求（仅检查条目级文件）"""
    if expense_category == '小交通':
        return True, ''

    query = ProjectFile.query.filter_by(project_id=project_id, file_type='invoice')
    if item_id:
        query = query.filter_by(item_id=item_id)
    else:
        query = query.filter(ProjectFile.item_id.isnot(None))
    has_invoice = query.first() is not None

    query = ProjectFile.query.filter_by(project_id=project_id, file_type='payment')
    if item_id:
        query = query.filter_by(item_id=item_id)
    else:
        query = query.filter(ProjectFile.item_id.isnot(None))
    has_payment = query.first() is not None

    missing = []
    if not has_invoice:
        missing.append('发票')
    if not has_payment:
        missing.append('支付记录')

    if missing:
        return False, f'缺少: {", ".join(missing)}'
    return True, ''


@project_bp.route('/<int:project_id>/items', methods=['POST'])
@login_required
def add_item(project_id):
    """添加备案条目"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        return jsonify({'error': '无权操作'}), 403

    try:
        expense_category = request.form.get('expense_category', '').strip()
        content = expense_category  # 报账内容即费用类别
        teacher = request.form.get('teacher', '').strip()
        reimbursement_date = request.form.get('reimbursement_date', '').strip()
        amount = request.form.get('amount', type=float)
        applicant = request.form.get('applicant', '').strip()
        student_id = request.form.get('student_id', '').strip()
        student_class = request.form.get('student_class', '').strip()

        if not all([expense_category, teacher, reimbursement_date, amount, applicant, student_id, student_class]):
            return jsonify({'error': '请填写所有必填字段'}), 400

        # 大交通费专属字段
        direction = None
        origin = None
        destination = None
        transport_mode = None
        if expense_category == '大交通费':
            direction = request.form.get('direction', '').strip() or None
            origin = request.form.get('origin', '').strip() or None
            destination = request.form.get('destination', '').strip() or None
            transport_mode = request.form.get('transport_mode', '').strip() or None

        item = ReimbursementItem(
            project_id=project_id,
            teacher=teacher,
            reimbursement_date=reimbursement_date,
            expense_category=expense_category,
            content=content,
            amount=amount,
            applicant=applicant,
            student_id=student_id,
            student_class=student_class,
            direction=direction,
            origin=origin,
            destination=destination,
            transport_mode=transport_mode
        )
        db.session.add(item)
        db.session.commit()

        return jsonify({
            'success': True,
            'item': item.to_dict(),
            'message': '条目已添加'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'添加失败: {str(e)}'}), 500


@project_bp.route('/<int:project_id>/items/<int:item_id>', methods=['PUT', 'POST'])
@login_required
def update_item(project_id, item_id):
    """更新备案条目"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        return jsonify({'error': '无权操作'}), 403

    item = ReimbursementItem.query.get_or_404(item_id)
    if item.project_id != project_id:
        return jsonify({'error': '条目不属于该项目'}), 400

    try:
        expense_category = request.form.get('expense_category', '').strip()
        item.content = expense_category  # 报账内容即费用类别
        item.expense_category = expense_category
        item.teacher = request.form.get('teacher', '').strip()
        item.reimbursement_date = request.form.get('reimbursement_date', '').strip()
        item.amount = request.form.get('amount', type=float)
        item.applicant = request.form.get('applicant', '').strip()
        item.student_id = request.form.get('student_id', '').strip()
        item.student_class = request.form.get('student_class', '').strip()

        # 大交通费专属字段
        if expense_category == '大交通费':
            item.direction = request.form.get('direction', '').strip() or None
            item.origin = request.form.get('origin', '').strip() or None
            item.destination = request.form.get('destination', '').strip() or None
            item.transport_mode = request.form.get('transport_mode', '').strip() or None
        else:
            item.direction = None
            item.origin = None
            item.destination = None
            item.transport_mode = None

        db.session.commit()
        return jsonify({
            'success': True,
            'item': item.to_dict(),
            'message': '条目已更新'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'更新失败: {str(e)}'}), 500


@project_bp.route('/<int:project_id>/items/<int:item_id>', methods=['DELETE'])
@login_required
def delete_item(project_id, item_id):
    """删除备案条目"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        return jsonify({'error': '无权操作'}), 403

    item = ReimbursementItem.query.get_or_404(item_id)
    if item.project_id != project_id:
        return jsonify({'error': '条目不属于该项目'}), 400

    try:
        # 删除关联的文件记录
        ProjectFile.query.filter_by(project_id=project_id, item_id=item_id).delete()
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '条目已删除'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'删除失败: {str(e)}'}), 500


@project_bp.route('/<int:project_id>/files/upload', methods=['POST'])
@login_required
def upload_file(project_id):
    """上传文件"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        return jsonify({'error': '无权操作'}), 403

    file = request.files.get('file')
    file_type = request.form.get('file_type', 'support')
    item_id = request.form.get('item_id', type=int) or None

    if not file:
        return jsonify({'error': '请选择文件'}), 400

    result = save_uploaded_file(file, project_id, file_type, current_user.id, item_id)
    if result:
        return jsonify({
            'success': True,
            'file': result.to_dict(),
            'message': '文件上传成功'
        })
    return jsonify({'error': '文件类型不允许'}), 400


@project_bp.route('/<int:project_id>/files/<int:file_id>/delete', methods=['POST'])
@login_required
def delete_file(project_id, file_id):
    """删除文件"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        if request.is_json:
            return jsonify({'error': '无权操作'}), 403
        flash('无权操作', 'danger')
        return redirect(url_for('project.dashboard'))

    file = ProjectFile.query.get_or_404(file_id)
    if file.project_id != project_id:
        if request.is_json:
            return jsonify({'error': '文件不属于该项目'}), 400
        flash('文件不属于该项目', 'danger')
        return redirect(url_for('project.project_detail', project_id=project_id))

    try:
        # 删除物理文件
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project_id), file.stored_filename)
        if os.path.exists(file_path):
            os.remove(file_path)

        db.session.delete(file)
        db.session.commit()

        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': '文件已删除'})
        flash('文件已删除', 'success')
    except Exception as e:
        db.session.rollback()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': f'删除失败: {str(e)}'}), 500
        flash(f'删除失败: {str(e)}', 'danger')

    return redirect(url_for('project.project_detail', project_id=project_id))


@project_bp.route('/<int:project_id>/files/<int:file_id>/download')
@login_required
def download_file(project_id, file_id):
    """下载文件"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))

    file = ProjectFile.query.get_or_404(file_id)
    if file.project_id != project_id:
        flash('文件不属于该项目', 'danger')
        return redirect(url_for('project.project_detail', project_id=project_id))

    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project_id))
    return send_from_directory(upload_dir, file.stored_filename,
                               download_name=file.original_filename, as_attachment=True)


@project_bp.route('/<int:project_id>/files/<int:file_id>/preview')
@login_required
def preview_file(project_id, file_id):
    """预览文件"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        return jsonify({'error': '无权访问'}), 403

    file = ProjectFile.query.get_or_404(file_id)
    if file.project_id != project_id:
        return jsonify({'error': '文件不属于该项目'}), 400

    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project_id))
    file_path = os.path.join(upload_dir, file.stored_filename)

    if not os.path.exists(file_path):
        return jsonify({'error': '文件不存在'}), 404

    ext = os.path.splitext(file.stored_filename)[1].lower()

    if ext in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
        mimetype = f'image/{ext[1:]}'
        if ext == '.jpg':
            mimetype = 'image/jpeg'
        return send_file(file_path, mimetype=mimetype)
    elif ext == '.pdf':
        return send_file(file_path, mimetype='application/pdf')
    else:
        # 其他类型直接下载
        return send_file(file_path, as_attachment=True, download_name=file.original_filename)


@project_bp.route('/<int:project_id>/files/check')
@login_required
def check_files(project_id):
    """检查文件状态"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        return jsonify({'error': '无权访问'}), 403

    item_id = request.args.get('item_id', type=int)
    expense_category = request.args.get('expense_category', '')

    invoice_query = ProjectFile.query.filter_by(project_id=project_id, file_type='invoice')
    payment_query = ProjectFile.query.filter_by(project_id=project_id, file_type='payment')
    if item_id is not None:
        invoice_query = invoice_query.filter_by(item_id=item_id)
        payment_query = payment_query.filter_by(item_id=item_id)
    else:
        invoice_query = invoice_query.filter(ProjectFile.item_id.isnot(None))
        payment_query = payment_query.filter(ProjectFile.item_id.isnot(None))

    invoice_files = invoice_query.order_by(ProjectFile.created_at.desc()).all()
    payment_files = payment_query.order_by(ProjectFile.created_at.desc()).all()

    ok, msg = check_item_file_requirement(project_id, expense_category, item_id)
    return jsonify({
        'ok': ok,
        'message': msg,
        'has_invoice': len(invoice_files) > 0,
        'has_payment': len(payment_files) > 0,
        'invoice_files': [f.to_dict() for f in invoice_files],
        'payment_files': [f.to_dict() for f in payment_files]
    })


@project_bp.route('/<int:project_id>/team/add', methods=['POST'])
@login_required
def add_team_member(project_id):
    """添加团队成员"""
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id and not current_user.is_admin():
        return jsonify({'error': '只有项目创建者或管理员可以管理团队'}), 403

    # 检查团队成员数量
    current_member_count = project.team_members.count()
    if current_member_count >= 3:
        return jsonify({'error': '团队成员已满（最多3人）'}), 400

    username = request.form.get('username', '').strip()
    user_id = request.form.get('user_id', type=int)
    
    user = None
    if user_id:
        user = User.query.get(user_id)
    elif username:
        user = User.query.filter_by(username=username).first()
    
    if not user:
        return jsonify({'error': '用户不存在'}), 404

    if user.id == project.owner_id:
        return jsonify({'error': '不能添加自己为团队成员'}), 400

    existing = TeamMember.query.filter_by(project_id=project.id, user_id=user.id).first()
    if existing:
        return jsonify({'error': '该用户已是团队成员'}), 400

    member = TeamMember(project_id=project.id, user_id=user.id)
    db.session.add(member)
    db.session.commit()
    return jsonify({'success': True, 'message': f'已添加团队成员: {user.real_name or user.username}'})


@project_bp.route('/<int:project_id>/team/<int:member_id>/remove', methods=['POST'])
@login_required
def remove_team_member(project_id, member_id):
    """移除团队成员"""
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id and not current_user.is_admin():
        return jsonify({'error': '只有项目创建者或管理员可以管理团队'}), 403

    member = TeamMember.query.get_or_404(member_id)
    if member.project_id != project_id:
        return jsonify({'error': '成员不属于该项目'}), 400

    db.session.delete(member)
    db.session.commit()
    return jsonify({'success': True, 'message': '已移除团队成员'})


def generate_reimbursement_excel(project, items):
    """生成备案表 Excel 文件（仅数据行，无标题/合计）"""
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '备案表'

    # 样式定义 - 全部黑色文字，无颜色填充
    header_font = Font(name='微软雅黑', size=11, bold=True, color='000000')
    data_font = Font(name='微软雅黑', size=10, color='000000')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 根据内容自动计算列宽
    headers = ['指导老师', '报账时间', '报账内容', '支出金额', '报账人', '学号', '班级', '对应比赛名称和时间']
    col_widths = [len(h) * 2.2 for h in headers]
    competition_text = project.competition_name or ''
    if project.competition_date:
        competition_text = f"{competition_text}\n{project.competition_date}"
    for item in items:
        content_display = item.expense_category or ''
        if item.expense_category == '大交通费':
            parts = []
            if item.direction:
                parts.append(item.direction)
            if item.origin and item.destination:
                parts.append(f"{item.origin}-{item.destination}")
            if item.transport_mode:
                parts.append(item.transport_mode)
            if parts:
                content_display = f"大交通{' '.join(parts)}"
        row_values = [
            item.teacher or '',
            item.reimbursement_date or '',
            content_display,
            f"¥{item.amount:.2f}" if item.amount else '',
            item.applicant or '',
            item.student_id or '',
            item.student_class or '',
            competition_text
        ]
        for i, val in enumerate(row_values):
            char_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
            needed = char_len * 1.1 + 2
            if needed > col_widths[i]:
                col_widths[i] = needed
    for i in range(len(col_widths)):
        col_widths[i] = max(8, min(col_widths[i], 50))
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 表头行（第1行）
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # 数据行
    first_data_row = 2
    last_data_row = len(items) + 1
    for row_idx, item in enumerate(items, 2):
        competition_text = project.competition_name or ''
        if project.competition_date:
            competition_text = f"{competition_text}\n{project.competition_date}"
        row_data = [
            item.teacher,
            item.reimbursement_date,
            item.expense_category,
            item.amount,
            item.applicant,
            item.student_id,
            item.student_class,
            competition_text
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align
            if col_idx == 4:
                cell.number_format = '¥#,##0.00'
        ws.row_dimensions[row_idx].height = 24

    # 合并比赛名称列（第8列）
    if len(items) > 1 and project.competition_name:
        ws.merge_cells(start_row=first_data_row, start_column=8,
                       end_row=last_data_row, end_column=8)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@project_bp.route('/<int:project_id>/export/excel')
@login_required
def export_excel(project_id):
    """导出备案表为 Excel 文件"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))

    if not HAS_OPENPYXL:
        flash('服务器缺少 openpyxl 库，无法生成 Excel 文件', 'danger')
        return redirect(url_for('project.project_detail', project_id=project_id))

    items = project.reimbursement_items.order_by(ReimbursementItem.created_at).all()
    if not items:
        flash('没有备案条目可导出', 'warning')
        return redirect(url_for('project.project_detail', project_id=project_id))

    output = generate_reimbursement_excel(project, items)
    if output is None:
        flash('生成备案表失败', 'danger')
        return redirect(url_for('project.project_detail', project_id=project_id))

    filename = f'备案表_{project.title}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@project_bp.route('/<int:project_id>/export/html')
@login_required
def export_html(project_id):
    """预览/导出备案表为 HTML 格式"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))

    items = project.reimbursement_items.order_by(ReimbursementItem.created_at).all()
    total_amount = sum(item.amount for item in items)

    return render_template('reimbursement_table.html',
                           project=project, items=items, total_amount=total_amount,
                           now=datetime.now)


def convert_to_image(file_path, output_dir, base_name):
    """将文件转换为PNG图片，支持PDF、Word、图片等格式
    
    返回: (成功标志, 生成的图片文件路径列表)
    """
    ext = os.path.splitext(file_path)[1].lower()
    generated_images = []

    try:
        if ext in ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'):
            # 已经是图片格式，直接复制并转换为PNG
            img = Image.open(file_path).convert('RGB')
            out_path = os.path.join(output_dir, f"{base_name}.png")
            img.save(out_path, 'PNG')
            generated_images.append(out_path)
            return True, generated_images

        elif ext == '.pdf':
            # PDF转图片
            if not HAS_PYMUPDF:
                return False, []
            pdf_doc = fitz.open(file_path)
            for page_num in range(len(pdf_doc)):
                page = pdf_doc[page_num]
                pix = page.get_pixmap(dpi=200)
                img_data = pix.tobytes("png")
                page_name = f"{base_name}_p{page_num + 1}.png"
                out_path = os.path.join(output_dir, page_name)
                with open(out_path, 'wb') as f:
                    f.write(img_data)
                generated_images.append(out_path)
            pdf_doc.close()
            return True, generated_images

        elif ext in ('.doc', '.docx'):
            # Word文档转图片 - 使用libreoffice转为PDF再转图片
            try:
                import subprocess
                temp_pdf = os.path.join(output_dir, f"{base_name}_temp.pdf")
                result = subprocess.run(
                    ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', output_dir, file_path],
                    capture_output=True, timeout=30
                )
                if result.returncode == 0:
                    pdf_name = os.path.splitext(os.path.basename(file_path))[0] + '.pdf'
                    pdf_path = os.path.join(output_dir, pdf_name)
                    if os.path.exists(pdf_path):
                        pdf_doc = fitz.open(pdf_path)
                        for page_num in range(len(pdf_doc)):
                            page = pdf_doc[page_num]
                            pix = page.get_pixmap(dpi=200)
                            img_data = pix.tobytes("png")
                            page_name = f"{base_name}_p{page_num + 1}.png"
                            out_path = os.path.join(output_dir, page_name)
                            with open(out_path, 'wb') as f:
                                f.write(img_data)
                            generated_images.append(out_path)
                        pdf_doc.close()
                        os.remove(pdf_path)
                        return True, generated_images
                return False, []
            except Exception:
                return False, []

        else:
            return False, []

    except Exception as e:
        print(f"转换文件失败 {file_path}: {e}")
        return False, []


def stitch_images_to_a4(image_paths, output_dir, base_name='支付记录_拼接图'):
    """将多张图片拼接为一张A4大小的图片（N×M网格排列）
    
    A4尺寸: 2480×3508 pixels @ 300dpi
    图片按N×M网格排列，每张图片保持原始宽高比，缩放到适合网格单元格大小
    所有图片拼接为一张A4图片，可留白不覆盖
    """
    if not image_paths:
        return False
    
    A4_WIDTH = 2480
    A4_HEIGHT = 3508
    
    try:
        n = len(image_paths)
        if n == 0:
            return False
        
        # 计算网格布局：尽量接近正方形
        import math
        cols = math.ceil(math.sqrt(n))
        rows = math.ceil(n / cols)
        
        # 每个网格单元格的尺寸（含间距）
        margin = 20  # 间距像素
        cell_w = (A4_WIDTH - margin * (cols + 1)) // cols
        cell_h = (A4_HEIGHT - margin * (rows + 1)) // rows
        
        # 创建A4画布（白色背景）
        canvas = Image.new('RGB', (A4_WIDTH, A4_HEIGHT), (255, 255, 255))
        
        for idx, img_path in enumerate(image_paths):
            row = idx // cols
            col = idx % cols
            
            img = Image.open(img_path).convert('RGB')
            
            # 按单元格尺寸等比缩放（保持宽高比，内切）
            img_w, img_h = img.size
            scale = min(cell_w / img_w, cell_h / img_h)
            new_w = int(img_w * scale)
            new_h = int(img_h * scale)
            img = img.resize((new_w, new_h), Image.LANCZOS)
            
            # 计算在单元格中的居中位置
            x = margin + col * (cell_w + margin) + (cell_w - new_w) // 2
            y = margin + row * (cell_h + margin) + (cell_h - new_h) // 2
            
            canvas.paste(img, (x, y))
        
        # 输出一张A4图片
        out_path = os.path.join(output_dir, f'{base_name}.png')
        canvas.save(out_path, 'PNG')
        return True
    except Exception as e:
        print(f"拼接图片失败: {e}")
        return False


@project_bp.route('/<int:project_id>/export/zip')
@login_required
def export_zip(project_id):
    """导出ZIP压缩包：备案表 + 发票/支付记录/辅助材料三个文件夹（文件转为图片）"""
    project = Project.query.get_or_404(project_id)
    if not check_project_access(project):
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))

    items = project.reimbursement_items.order_by(ReimbursementItem.created_at).all()

    # 创建临时目录
    temp_dir = tempfile.mkdtemp()
    try:
        # 1. 生成备案表Excel
        excel_output = generate_reimbursement_excel(project, items)
        if excel_output:
            excel_path = os.path.join(temp_dir, f'备案表_{project.title}.xlsx')
            with open(excel_path, 'wb') as f:
                f.write(excel_output.getvalue())

        # 2. 收集所有文件，按类型分类
        all_files = ProjectFile.query.filter_by(project_id=project_id).all()

        # 创建三个文件夹
        invoice_dir = os.path.join(temp_dir, '发票')
        payment_dir = os.path.join(temp_dir, '支付记录')
        support_dir = os.path.join(temp_dir, '辅助材料')
        os.makedirs(invoice_dir, exist_ok=True)
        os.makedirs(payment_dir, exist_ok=True)
        os.makedirs(support_dir, exist_ok=True)

        # 上传目录
        upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project_id))

        # 文件重命名计数器（避免同名冲突）
        name_counters = {'invoice': {}, 'payment': {}, 'support': {}}

        # 支付记录图片收集（后续需要拼接为A4图片）
        # 使用tempfile创建独立临时目录，避免被ZIP打包
        payment_temp_dir = tempfile.mkdtemp()
        payment_images = []

        for pf in all_files:
            file_path = os.path.join(upload_dir, pf.stored_filename)
            if not os.path.exists(file_path):
                continue

            # 确定目标文件夹
            if pf.file_type == 'invoice':
                target_dir = invoice_dir
                type_key = 'invoice'
            elif pf.file_type == 'payment':
                # 支付记录：先转换到独立临时目录，收集图片路径
                target_dir = payment_temp_dir
                type_key = 'payment'
            else:
                target_dir = support_dir
                type_key = 'support'

            # 确定文件命名
            # 如果有关联条目，使用条目信息命名；否则使用原文件名
            if pf.item_id:
                item = ReimbursementItem.query.get(pf.item_id)
                if item:
                    base_name = f"{item.applicant}_{item.expense_category}_{item.reimbursement_date}"
                else:
                    base_name = os.path.splitext(pf.original_filename)[0]
            else:
                base_name = os.path.splitext(pf.original_filename)[0]

            # 处理重名
            counter = name_counters[type_key].get(base_name, 0) + 1
            name_counters[type_key][base_name] = counter
            if counter > 1:
                unique_base = f"{base_name}_{counter}"
            else:
                unique_base = base_name

            # 转换为图片
            success, images = convert_to_image(file_path, target_dir, unique_base)
            if success:
                if pf.file_type == 'payment':
                    payment_images.extend(images)
            else:
                # 转换失败，直接复制原文件（作为后备）
                try:
                    ext = os.path.splitext(pf.stored_filename)[1]
                    dest_path = os.path.join(target_dir, f"{unique_base}{ext}")
                    shutil.copy2(file_path, dest_path)
                    if pf.file_type == 'payment':
                        # 非图片文件无法拼接，直接复制到支付记录文件夹
                        fallback_dest = os.path.join(payment_dir, f"{unique_base}{ext}")
                        shutil.copy2(file_path, fallback_dest)
                except Exception:
                    pass

        # 将支付记录的所有图片拼接为A4图片（直接输出到payment_dir）
        if payment_images:
            stitch_images_to_a4(payment_images, payment_dir)

        # 清理支付记录临时目录
        shutil.rmtree(payment_temp_dir, ignore_errors=True)

        # 3. 打包为ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zf.write(file_path, arcname)

        zip_buffer.seek(0)

        filename = f'报销材料_{project.title}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'导出失败: {str(e)}', 'danger')
        return redirect(url_for('project.project_detail', project_id=project_id))
    finally:
        # 清理临时目录
        shutil.rmtree(temp_dir, ignore_errors=True)


# ==================== 管理员功能 ====================


@project_bp.route('/admin/users')
@login_required
def admin_users():
    if not current_user.is_admin():
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin_users.html', users=users)


@project_bp.route('/admin/users/<int:user_id>/toggle-role', methods=['POST'])
@login_required
def toggle_user_role(user_id):
    if not current_user.is_admin():
        return jsonify({'error': '无权操作'}), 403

    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({'error': '不能修改自己的角色'}), 400

    user.role = 'admin' if user.role == 'user' else 'user'
    db.session.commit()
    return jsonify({'success': True, 'role': user.role})


@project_bp.route('/admin/projects')
@login_required
def admin_projects():
    if not current_user.is_admin():
        flash('无权访问', 'danger')
        return redirect(url_for('project.dashboard'))
    projects = Project.query.order_by(Project.updated_at.desc()).all()
    return render_template('admin_projects.html', projects=projects)


@project_bp.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    """管理员删除用户（同时删除该用户创建的所有项目及关联数据）"""
    if not current_user.is_admin():
        return jsonify({'error': '无权操作'}), 403

    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({'error': '不能删除自己'}), 400

    if user.is_admin():
        admin_count = User.query.filter_by(role='admin').count()
        if admin_count <= 1:
            return jsonify({'error': '至少保留一名管理员'}), 400

    try:
        # 1. 删除该用户创建的项目（含关联的条目、文件、团队成员记录）
        user_projects = Project.query.filter_by(owner_id=user.id).all()
        for project in user_projects:
            # 删除项目上传的文件
            upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], str(project.id))
            if os.path.exists(upload_dir):
                shutil.rmtree(upload_dir)

            # 删除项目的所有文件记录
            ProjectFile.query.filter_by(project_id=project.id).delete()
            # 删除项目的所有条目
            ReimbursementItem.query.filter_by(project_id=project.id).delete()
            # 删除项目的所有团队成员记录
            TeamMember.query.filter_by(project_id=project.id).delete()
            # 删除项目本身
            db.session.delete(project)

        # 2. 删除用户作为团队成员参与的记录
        TeamMember.query.filter_by(user_id=user.id).delete()

        # 3. 删除用户的验证码记录
        EmailVerificationCode.query.filter_by(email=user.email).delete()

        # 4. 删除用户
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True, 'message': '用户已删除'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'删除失败: {str(e)}'}), 500
